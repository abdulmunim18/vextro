"""Generate exportable SME competitor-intelligence reports."""

from io import BytesIO

from app.schemas.sme import CompetitorIntelligenceResponse


def _safe_spreadsheet_text(value: str | None) -> str:
    """Prevent spreadsheet formula execution in exported text cells."""

    text = value or ""

    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"

    return text


def build_competitor_xlsx(
    *,
    organization_name: str,
    intelligence: CompetitorIntelligenceResponse,
) -> bytes:
    """Create a formatted Excel competitor report in memory."""

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["VEXTRO Competitor Intelligence Report"])
    summary_sheet.append(["Organization", organization_name])
    summary_sheet.append(
        ["Generated at", intelligence.generated_at.isoformat()]
    )
    summary_sheet.append([])

    summary = intelligence.summary
    summary_rows = [
        ("Tracked competitors", summary.tracked_competitors),
        ("Tracked products", summary.tracked_products),
        ("Average price gap", summary.average_price_gap),
        ("Products at risk", summary.products_at_risk),
        (
            "Estimated average market share (%)",
            summary.estimated_average_market_share_percentage,
        ),
        ("Estimation note", summary.estimation_note),
    ]

    for label, value in summary_rows:
        summary_sheet.append([label, value])

    summary_sheet["A1"].font = Font(
        bold=True,
        size=16,
        color="FFFFFF",
    )
    summary_sheet["A1"].fill = PatternFill(
        "solid",
        fgColor="1D4ED8",
    )
    summary_sheet.merge_cells("A1:B1")
    summary_sheet.column_dimensions["A"].width = 40
    summary_sheet.column_dimensions["B"].width = 75
    summary_sheet.freeze_panes = "A5"

    detail_sheet = workbook.create_sheet("Competitor Details")
    headers = [
        "Product",
        "Platform",
        "Seller",
        "Own Price",
        "Competitor Price",
        "Gap",
        "Gap %",
        "Position",
        "Risk",
        "Estimated Own Share %",
        "Risk Reasons",
    ]
    detail_sheet.append(headers)

    for item in intelligence.items:
        detail_sheet.append(
            [
                _safe_spreadsheet_text(item.own_product_name),
                _safe_spreadsheet_text(item.platform_name),
                _safe_spreadsheet_text(item.seller_name),
                item.own_price,
                item.competitor_price,
                item.price_gap,
                item.price_gap_percentage,
                item.price_position,
                item.risk_level,
                item.estimated_own_market_share_percentage,
                _safe_spreadsheet_text("; ".join(item.risk_reasons)),
            ]
        )

    header_fill = PatternFill("solid", fgColor="0F172A")

    for cell in detail_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    widths = [
        30,
        16,
        24,
        16,
        18,
        16,
        12,
        20,
        12,
        22,
        52,
    ]

    for index, width in enumerate(widths, start=1):
        detail_sheet.column_dimensions[
            detail_sheet.cell(row=1, column=index).column_letter
        ].width = width

    for row in detail_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    detail_sheet.freeze_panes = "A2"
    detail_sheet.auto_filter.ref = detail_sheet.dimensions
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_competitor_pdf(
    *,
    organization_name: str,
    intelligence: CompetitorIntelligenceResponse,
) -> bytes:
    """Create a readable landscape PDF competitor report in memory."""

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    stream = BytesIO()
    document = SimpleDocTemplate(
        stream,
        pagesize=landscape(A4),
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title="VEXTRO Competitor Intelligence Report",
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("VEXTRO Competitor Intelligence", styles["Title"]),
        Paragraph(
            f"Organization: {organization_name}",
            styles["Heading2"],
        ),
        Paragraph(
            f"Generated: {intelligence.generated_at.isoformat()}",
            styles["BodyText"],
        ),
        Spacer(1, 5 * mm),
    ]
    summary = intelligence.summary
    summary_data = [
        ["Tracked competitors", "Tracked products", "At risk", "Avg gap"],
        [
            str(summary.tracked_competitors),
            str(summary.tracked_products),
            str(summary.products_at_risk),
            str(summary.average_price_gap or "N/A"),
        ],
    ]
    summary_table = Table(
        summary_data,
        colWidths=[62 * mm] * 4,
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("PADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 6 * mm)])
    detail_data = [
        [
            "Product",
            "Platform / Seller",
            "Own",
            "Competitor",
            "Gap %",
            "Risk",
            "Est. share %",
        ]
    ]

    for item in intelligence.items:
        detail_data.append(
            [
                Paragraph(item.own_product_name, styles["BodyText"]),
                Paragraph(
                    f"{item.platform_name}<br/>{item.seller_name or 'N/A'}",
                    styles["BodyText"],
                ),
                str(item.own_price or "N/A"),
                str(item.competitor_price),
                str(item.price_gap_percentage or "N/A"),
                item.risk_level.upper(),
                str(item.estimated_own_market_share_percentage or "N/A"),
            ]
        )

    detail_table = Table(
        detail_data,
        repeatRows=1,
        colWidths=[55 * mm, 52 * mm, 29 * mm, 34 * mm, 23 * mm, 22 * mm, 28 * mm],
    )
    detail_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (2, 1), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.extend(
        [
            detail_table,
            Spacer(1, 5 * mm),
            Paragraph(summary.estimation_note, styles["Italic"]),
        ]
    )
    document.build(story)
    return stream.getvalue()
